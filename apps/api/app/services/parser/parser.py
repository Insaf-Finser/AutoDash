from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.parser.models import WorkbookMetadata
from app.services.parser.models import WorksheetMetadata
from zipfile import BadZipFile
from fastapi import HTTPException, status


class WorkbookParser:
    """
    Extract workbook metadata.
    """

    def extract_metadata(
        self,
        data: bytes,
    ) -> WorkbookMetadata:

        try:
            workbook = load_workbook(
                filename=BytesIO(data),
                read_only=True,
                data_only=True,
            )
        except BadZipFile:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid or corrupted Excel (.xlsx) file.",
            )

        worksheets: list[WorksheetMetadata] = []

        for sheet in workbook.worksheets:
            worksheets.append(
                WorksheetMetadata(
                    name=sheet.title,
                    row_count=sheet.max_row,
                    column_count=sheet.max_column,
                )
            )

        workbook.close()

        return WorkbookMetadata(
            sheet_count=len(worksheets),
            worksheets=worksheets,
        )